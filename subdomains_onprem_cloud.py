#!/usr/bin/env python3
import socket
import ssl
import subprocess
import requests
import dns.resolver
from openpyxl import Workbook
from openpyxl.styles import PatternFill

SHODAN_API_KEY = "${SHODAN_API_KEY}"
DOMINIOS_TXT = "dominios.txt"
CLOUD_KEYWORDS = ["amazon", "aws", "google", "cloud", "azure", "digitalocean", "oracle", "microsoft", "linode", "gcp", "cloudflare", "fastly", "alibaba"]

BANNER = r"""
 ____  _   _ ____  ____   ___  __  __    _    ___ _   _ ____
/ ___|| | | | __ )|  _ \ / _ \|  \/  |  / \  |_ _| \ | / ___|
\___ \| | | |  _ \| | | | | | | |\/| | / _ \  | ||  \| \___ \
 ___) | |_| | |_) | |_| | |_| | |  | |/ ___ \ | || |\  |___) |
|____/ \___/|____/|____/ \___/|_|  |_/_/   \_\___|_| \_|____/
"""


def resolve_dns(domain, record_type):
    try:
        return [r.to_text() for r in dns.resolver.resolve(domain, record_type, lifetime=3)]
    except Exception:
        return []


def get_ipinfo(ip):
    try:
        res = requests.get(f"https://ipinfo.io/{ip}/json", timeout=5)
        return res.json() if res.status_code == 200 else {}
    except Exception:
        return {}


def classify(info):
    org = info.get("org", "").lower()
    return "Cloud" if any(k in org for k in CLOUD_KEYWORDS) else ("On-premise" if org else "Unknown")


def load_domains(filename):
    with open(filename, "r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip()]


def export_to_excel(results, filename="resultados_dominios.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.append(["Dominio", "IP", "Organización", "Ubicación"])
    fills = {
        "Cloud": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "On-premise": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "Unknown": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }
    for r in results:
        ws.append([r["domain"], r["ip"], r["org"], r["location"]])
        for cell in ws[ws.max_row]:
            cell.fill = fills.get(r["location"], fills["Unknown"])
    wb.save(filename)


def main():
    print(BANNER)
    print("[DISCLAIMER] Uso autorizado únicamente.")
    print("Autor: met4ll0f | https://github.com/aka-met4ll0f")
    domains = load_domains(DOMINIOS_TXT)
    results = []
    for domain in domains:
        a = resolve_dns(domain, "A")
        ip = a[0] if a else "N/A"
        info = get_ipinfo(ip) if ip != "N/A" else {}
        results.append({"domain": domain, "ip": ip, "org": info.get("org", "N/A"), "location": classify(info)})
    export_to_excel(results)


if __name__ == "__main__":
    main()
